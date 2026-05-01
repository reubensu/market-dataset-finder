from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def save_dataframe_workbook(df: pd.DataFrame, filename: Path, sheet_name: str) -> None:
    try:
        df.to_excel(filename, index=False, sheet_name=sheet_name)
        format_workbook(filename)
    except PermissionError as exc:
        raise PermissionError(
            f"Could not write {filename}. Close the Excel file if it is open and try again."
        ) from exc


def save_multi_sheet_workbook(sheets: dict[str, pd.DataFrame], filename: Path) -> None:
    try:
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                safe_name = sheet_name[:31]
                df.to_excel(writer, sheet_name=safe_name, index=False)
        format_workbook(filename)
    except PermissionError as exc:
        raise PermissionError(
            f"Could not write {filename}. Close the Excel file if it is open and try again."
        ) from exc


def format_workbook(filename: Path) -> None:
    workbook = load_workbook(filename)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for column_cells in sheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 55)

    workbook.save(filename)
