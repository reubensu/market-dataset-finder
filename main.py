from huggingface_hub import HfApi
import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


SEARCH_TERMS = [
    "ecommerce",
    "shopping",
    "product reviews",
    "consumer behaviour",
    "market research",
    "amazon reviews",
    "online retail",
    "customer reviews",
    "product sentiment",
    "marketplace",
    "shopee product reviews",
]


POSITIVE_KEYWORDS = {
    "amazon": 35,
    "review": 35,
    "reviews": 35,
    "product": 30,
    "products": 30,
    "ecommerce": 30,
    "e-commerce": 30,
    "shopping": 25,
    "retail": 25,
    "customer": 20,
    "sentiment": 20,
    "marketplace": 20,
    "rating": 15,
    "ratings": 15,
    "consumer": 15,
    "shopee": 30,
}


NEGATIVE_KEYWORDS = {
    "audio": -40,
    "image": -25,
    "video": -25,
    "chatbot": -20,
    "faq": -15,
    "games": -15,
    "multilingual": -10,
}


DECISION_RANKS = {
    "strong candidate": 1,
    "review manually": 2,
    "secondary": 3,
    "check licence": 4,
    "check access": 5,
    "low priority": 6,
}


def clean_text(value):
    if value is None:
        return ""

    return str(value).replace("\n", " ").strip()


def shorten_text(value, max_length=220):
    value = clean_text(value)

    if len(value) <= max_length:
        return value

    return value[:max_length].strip() + "..."


def join_list(value, max_items=8):
    if value is None:
        return ""

    if isinstance(value, list):
        return ", ".join(str(item) for item in value[:max_items])

    return str(value)


def get_card_data(dataset):
    card_data = getattr(dataset, "card_data", None)

    if card_data is None:
        return {}

    if isinstance(card_data, dict):
        return card_data

    if hasattr(card_data, "to_dict"):
        return card_data.to_dict()

    return {}


def get_card_field(dataset, field_name):
    card_data = get_card_data(dataset)
    value = card_data.get(field_name, "")

    return join_list(value)


def classify_dataset(text):
    text = text.lower()

    if "shopee" in text:
        return "shopee / marketplace"

    if "amazon" in text and "review" in text:
        return "amazon product reviews"

    if "product" in text and "review" in text:
        return "product reviews"

    if "sentiment" in text:
        return "sentiment analysis"

    if "ecommerce" in text or "e-commerce" in text or "retail" in text:
        return "e-commerce / retail"

    if "customer" in text:
        return "customer behaviour / support"

    if "image" in text:
        return "image / visual search"

    return "general / needs review"


def calculate_relevance(dataset_id, search_term, description, tags, downloads, likes):
    text = f"{dataset_id} {search_term} {description} {tags}".lower()

    keyword_score = 0

    for keyword, weight in POSITIVE_KEYWORDS.items():
        if keyword in text:
            keyword_score += weight

    for keyword, weight in NEGATIVE_KEYWORDS.items():
        if keyword in text:
            keyword_score += weight

    popularity_score = math.log10(downloads + 1) * 20
    like_score = math.log10(likes + 1) * 15

    return round(keyword_score + popularity_score + like_score, 2)


def calculate_metadata_quality(description, licence, language, task_categories, tags):
    score = 0

    if description:
        score += 20

    if licence:
        score += 20

    if language:
        score += 10

    if task_categories:
        score += 15

    if tags:
        score += 10

    return score


def infer_use_case(category):
    if category == "amazon product reviews":
        return "review analysis, sentiment, ratings, product categories"

    if category == "shopee / marketplace":
        return "marketplace research, product categories, regional e-commerce"

    if category == "product reviews":
        return "sentiment analysis, complaint mining, consumer preferences"

    if category == "e-commerce / retail":
        return "online retail analysis, catalogue or behaviour research"

    if category == "customer behaviour / support":
        return "customer-service pattern analysis"

    if category == "image / visual search":
        return "visual product search; not priority for text review analysis"

    return "manual review needed"


def create_decision(category, overall_score, downloads, licence, access):
    if access == "gated":
        return "check access"

    if not licence:
        return "check licence"

    if overall_score >= 175 and downloads >= 100:
        return "strong candidate"

    if "product reviews" in category or "amazon" in category or "shopee" in category:
        return "review manually"

    if "image" in category or "support" in category:
        return "secondary"

    return "low priority"


def create_reason(category, overall_score, downloads, licence, access):
    reasons = []

    if access == "gated":
        reasons.append("access may be restricted")

    if not licence:
        reasons.append("licence missing")

    if "product reviews" in category or "amazon" in category or "shopee" in category:
        reasons.append("matches product/e-commerce review focus")

    if downloads >= 100:
        reasons.append("has meaningful download activity")

    if overall_score >= 175:
        reasons.append("high overall score")

    if not reasons:
        reasons.append("weak relevance or limited metadata")

    return "; ".join(reasons)


def search_huggingface_datasets(search_term, limit=20):
    api = HfApi()

    datasets = api.list_datasets(
        search=search_term,
        limit=limit,
        full=True
    )

    results = []

    for dataset in datasets:
        dataset_id = dataset.id
        downloads = dataset.downloads or 0
        likes = dataset.likes or 0

        description = shorten_text(getattr(dataset, "description", ""))
        tags = join_list(getattr(dataset, "tags", []))
        licence = get_card_field(dataset, "license")
        language = get_card_field(dataset, "language")
        task_categories = get_card_field(dataset, "task_categories")
        access = "gated" if getattr(dataset, "gated", False) else "public"

        category = classify_dataset(
            f"{dataset_id} {search_term} {description} {tags} {task_categories}"
        )

        relevance_score = calculate_relevance(
            dataset_id=dataset_id,
            search_term=search_term,
            description=description,
            tags=tags,
            downloads=downloads,
            likes=likes
        )

        metadata_quality_score = calculate_metadata_quality(
            description=description,
            licence=licence,
            language=language,
            task_categories=task_categories,
            tags=tags
        )

        overall_score = round(relevance_score + (metadata_quality_score * 0.3), 2)

        decision = create_decision(
            category=category,
            overall_score=overall_score,
            downloads=downloads,
            licence=licence,
            access=access
        )

        decision_rank = DECISION_RANKS.get(decision, 99)

        reason = create_reason(
            category=category,
            overall_score=overall_score,
            downloads=downloads,
            licence=licence,
            access=access
        )

        results.append({
            "source": "hugging face",
            "search_term": search_term,
            "dataset_id": dataset_id,
            "category": category,
            "use_case": infer_use_case(category),
            "downloads": downloads,
            "likes": likes,
            "relevance_score": relevance_score,
            "metadata_quality_score": metadata_quality_score,
            "overall_score": overall_score,
            "decision_rank": decision_rank,
            "decision": decision,
            "reason": reason,
            "access": access,
            "licence": licence,
            "language": language,
            "task_categories": task_categories,
            "tags": tags,
            "description": description,
            "url": f"https://huggingface.co/datasets/{dataset_id}"
        })

    return results


def format_excel_file(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    sheet.title = "dataset shortlist"

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = min(max_length + 2, 45)

    workbook.save(filename)


def main():
    all_results = []

    for term in SEARCH_TERMS:
        print(f"searching for: {term}")
        results = search_huggingface_datasets(term)
        all_results.extend(results)

    df = pd.DataFrame(all_results)

    df = df.drop_duplicates(subset=["dataset_id"])

    df = df.sort_values(
        by=["decision_rank", "overall_score", "relevance_score", "downloads"],
        ascending=[True, False, False, False]
    )

    df.to_csv("dataset_results.csv", index=False)
    df.to_excel("dataset_results.xlsx", index=False)

    format_excel_file("dataset_results.xlsx")

    print(f"done. saved {len(df)} unique datasets.")
    print("files created: dataset_results.csv and dataset_results.xlsx")


if __name__ == "__main__":
    main()